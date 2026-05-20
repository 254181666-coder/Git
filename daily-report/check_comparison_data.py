
#!/usr/bin/env python3
"""
查看同比分析的原始数据
"""
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.database import query
import pandas as pd
from collections import Counter

STORE_NAME_MERGE = {
    '佳木斯': '佳木斯', '佳木斯店': '佳木斯',
    '安达': '安达', '安达店': '安达',
    '晨宇': '晨宇', '晨宇店': '晨宇',
    '鸡西': '鸡西', '鸡西店': '鸡西',
    '通辽': '通辽', '通辽店': '通辽',
    '松原一': '松原一', '松原一店': '松原一',
    '松原二': '松原二', '松原二店': '松原二',
    '榆树': '榆树', '榆树店': '榆树',
    '法库': '法库', '法库店': '法库',
    '锡盟': '锡盟', '锡盟店': '锡盟',
}

GB_SOURCES = {'抖音', '美团大众', '线下团购'}

def unify(n):
    if not n:
        return None
    return STORE_NAME_MERGE.get(str(n).strip(), str(n).strip())

def store_map():
    df = query("SELECT id, store_name FROM stores")
    return dict(zip(df['id'], df['store_name']))

sm = store_map()

print("=" * 80)
print("查看佳木斯和安达的对比数据")
print("=" * 80)

# 查看2025年Excel数据
from openpyxl import load_workbook
xlsx = load_workbook("/Users/ann/Desktop/25年团购内容.xlsx", data_only=True)

print("\n=== 2025年数据 ===")
for sheet_name in xlsx.sheetnames:
    ws = xlsx[sheet_name]
    print(f"\nSheet: {sheet_name}")
    
    last_year_data = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[6]:
            continue
        store = str(row[6]).strip()
        if store not in ['佳木斯', '安达']:
            continue
        
        if store not in last_year_data:
            last_year_data[store] = {'美团订单': 0, '抖音订单': 0, '总营收': 0, '订单数': 0}
        
        platform = sheet_name.replace('Sheet', '')
        price = row[2] if row[2] else 0
        if '美团' in sheet_name:
            last_year_data[store]['美团订单'] += 1
        if '抖音' in sheet_name:
            last_year_data[store]['抖音订单'] += 1
        
        last_year_data[store]['总营收'] += price
        last_year_data[store]['订单数'] += 1

    for store in ['佳木斯', '安达']:
        if store in last_year_data:
            data = last_year_data[store]
            print(f"\n{store}:")
            print(f"  美团订单: {data['美团订单']}")
            print(f"  抖音订单: {data['抖音订单']}")
            print(f"  总订单: {data['订单数']}")
            print(f"  总营收: {data['总营收']:,.0f}")

print("\n" + "=" * 80)
print("=== 2026年数据 ===")

# 加载2026年数据
od = query("""
    SELECT store_id, data_date, order_no, source_channel, actual_amount
    FROM order_detail
    WHERE data_date >= '2026-05-01' AND data_date <= '2026-05-11'
    AND order_type IN ('开房单', '点单')
    AND source_channel IN ('抖音', '美团大众', '线下团购')
""")

if not od.empty:
    od['门店'] = od['store_id'].map(sm).apply(unify)
    od = od[od['门店'].notna()]
    
    for store in ['佳木斯', '安达']:
        store_orders = od[od['门店'] == store]
        print(f"\n{store}:")
        print(f"  总订单: {len(store_orders)}")
        print(f"  总营收: {store_orders['actual_amount'].sum():,.0f}")
        
        print(f"  各平台:")
        for channel in ['抖音', '美团大众', '线下团购']:
            channel_orders = store_orders[store_orders['source_channel'] == channel]
            print(f"    {channel}: {len(channel_orders)} 单, {channel_orders['actual_amount'].sum():,.0f} 元")

