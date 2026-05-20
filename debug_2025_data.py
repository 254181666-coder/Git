
#!/usr/bin/env python3
"""
正确解析2025年Excel数据
"""
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.database import query
import pandas as pd
from openpyxl import load_workbook

print("=" * 80)
print("正确解析2025年团购数据")
print("=" * 80)

xlsx = load_workbook("/Users/ann/Desktop/25年团购内容.xlsx", data_only=True)

stores_data = {}

for sheet_name in xlsx.sheetnames:
    ws = xlsx[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    
    current_store = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        # 第6列是门店名
        store_cell = row[6]
        if store_cell and isinstance(store_cell, str) and len(store_cell.strip()) > 0 and not any(x in store_cell for x in ['团购', '套餐', 'Sheet']):
            current_store = store_cell.strip()
            if current_store not in stores_data:
                stores_data[current_store] = {'抖音订单':0, '美团订单':0, '抖音营收':0, '美团营收':0, '总订单':0, '总营收':0}
            continue
        
        # 有数据的行
        if current_store and row[1] and row[2]:  # 有套餐名和价格
            price_val = row[2]
            if isinstance(price_val, str) and ('价格' in price_val or not price_val):
                continue
            
            try:
                price = float(price_val)
            except:
                continue
            
            if '抖音' in sheet_name:
                stores_data[current_store]['抖音订单'] += 1
                stores_data[current_store]['抖音营收'] += price
            else:
                stores_data[current_store]['美团订单'] += 1
                stores_data[current_store]['美团营收'] += price
            
            stores_data[current_store]['总订单'] += 1
            stores_data[current_store]['总营收'] += price

# 打印所有门店
print("\n" + "=" * 80)
print("各门店2025年数据")
print("=" * 80)
for store in sorted(stores_data.keys()):
    data = stores_data[store]
    print(f"\n{store}:")
    print(f"  美团: {data['美团订单']} 单, {data['美团营收']:,.0f} 元")
    print(f"  抖音: {data['抖音订单']} 单, {data['抖音营收']:,.0f} 元")
    print(f"  总计: {data['总订单']} 单, {data['总营收']:,.0f} 元")

# 查看佳木斯和安达
print("\n" + "=" * 80)
print("佳木斯和安达详细")
print("=" * 80)
for store in ['佳木斯', '安达']:
    if store in stores_data:
        print(f"\n{store}:")
        print(f"  美团: {stores_data[store]['美团订单']} 单")
        print(f"  抖音: {stores_data[store]['抖音订单']} 单")
        print(f"  总计: {stores_data[store]['总订单']} 单, {stores_data[store]['总营收']:,.0f} 元")

# 加载2026年数据对比
print("\n" + "=" * 80)
print("2026年数据对比")
print("=" * 80)

STORE_NAME_MERGE = {
    '佳木斯': '佳木斯', '佳木斯店': '佳木斯',
    '安达': '安达', '安达店': '安达',
}

def unify(n):
    if not n:
        return None
    return STORE_NAME_MERGE.get(str(n).strip(), str(n).strip())

def store_map():
    df = query("SELECT id, store_name FROM stores")
    return dict(zip(df['id'], df['store_name']))

sm = store_map()

od = query("""
    SELECT store_id, data_date, order_no, source_channel, actual_amount
    FROM order_detail
    WHERE data_date >= '2026-05-01' AND data_date <= '2026-05-11'
    AND order_type IN ('开房单', '点单')
    AND source_channel IN ('抖音', '美团大众', '线下团购')
""")

if not od.empty:
    od['门店'] = od['store_id'].map(sm).apply(unify)
    od = od[od['门店'].notna()]
    
    for store in ['佳木斯', '安达']:
        store_orders = od[od['门店'] == store]
        print(f"\n{store} 2026:")
        print(f"  总订单: {len(store_orders)} 单")
        print(f"  总营收: {store_orders['actual_amount'].sum():,.0f} 元")

# 计算百分比
print("\n" + "=" * 80)
print("同比百分比计算示例")
print("=" * 80)

for store in ['佳木斯', '安达']:
    if store in stores_data and len(od) > 0:
        ly = stores_data[store]
        ty_store = od[od['门店'] == store]
        ty_orders = len(ty_store)
        ty_rev = ty_store['actual_amount'].sum()
        
        ly_orders = ly['总订单']
        ly_rev = ly['总营收']
        
        order_pct = (ty_orders - ly_orders) / ly_orders * 100 if ly_orders > 0 else 0
        rev_pct = (ty_rev - ly_rev) / ly_rev * 100 if ly_rev > 0 else 0
        
        print(f"\n{store}:")
        print(f"  订单: 2025年={ly_orders}, 2026年={ty_orders}")
        print(f"    变化: {order_pct:+.1f}%")
        
        print(f"  营收: 2025年={ly_rev:.0f}, 2026年={ty_rev:.0f}")
        print(f"    变化: {rev_pct:+.1f}%")

